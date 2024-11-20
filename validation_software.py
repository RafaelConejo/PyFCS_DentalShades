import tkinter as tk
from tkinter import ttk, simpledialog, messagebox
from PIL import Image, ImageTk
import os
import sys
import random
from openpyxl import Workbook, load_workbook

current_dir = os.path.dirname(__file__)
sys.path.append(current_dir)

# Create main window
root = tk.Tk()
root.geometry("1200x900") 
root.title("Color Selector")

def ask_user_name():
    """Asks the user to enter a name."""
    global user_name
    user_name = simpledialog.askstring("ID", "Enter user ID:")
    if not user_name:  # If the user doesn't enter anything
        messagebox.showerror("Error", "You must enter a name to continue.")
        ask_user_name()  # Ask for the name again


# Function to load the corresponding image from the 15 selected options
def update_image_from_selection():
    global image_id
    selected_label = selected_color.get()  # Get the selected label (e.g., "A1", "B2", etc.)
    
    # If no selection is made, clear the placeholder
    if not selected_label:
        placeholder_1.delete("all")  # Clear content of Image 1
        placeholder_1.create_text(50, 60, text=" ")  
        return
    
    # Determine the directory based on the selected option (tooth or reference)
    if image_source.get() == "tooth":
        directory = "Datasets/vita_tooth_test"  
    else:
        directory = "Datasets/reference_colors"  

    # Build the image path
    image_path = os.path.join(directory, f"{selected_label}.png")

    # Check if the image exists in the directory
    if os.path.exists(image_path):
        # Load the image and preserve transparency
        img = Image.open(image_path).convert("RGBA")
        img_resized = img.resize((77, 112), Image.Resampling.LANCZOS)  # Resize the image
        img_tk = ImageTk.PhotoImage(img_resized)  # Convert the image to Tkinter format

        # Clear the canvas before adding the new image
        placeholder_1.delete("all")
        placeholder_1.image = img_tk  # Keep the reference to avoid garbage collection
        image_id = placeholder_1.create_image(50, 60, image=img_tk)  # Draw the image on the canvas

        # Link click with move
        placeholder_1.bind("<ButtonPress-1>", on_image_press)  # When the user presses the mouse
        placeholder_1.bind("<B1-Motion>", on_image_drag)      # When the user drags the image
        placeholder_1.bind("<ButtonRelease-1>", on_image_release)
        
    else:
        # If the image is not found, show an error message
        messagebox.showerror("Error", f"Image not found: {image_path}")
        placeholder_1.delete("all")
        placeholder_1.create_text(50, 60, text="Image not found")

def on_image_press(event):
    global offset_x, offset_y
    # Guardamos las coordenadas iniciales donde el usuario hace clic (coordenadas globales)
    offset_x = event.x_root  # Coordenada X global
    offset_y = event.y_root  # Coordenada Y global

def on_image_drag(event):
    global offset_x, offset_y
    # Calculamos el desplazamiento de la imagen en relación con las coordenadas globales
    dx = event.x_root - offset_x
    dy = event.y_root - offset_y
    
    # Mover la imagen dentro del Canvas
    placeholder_1.move(image_id, dx, dy)
    
    # Actualizamos la posición de la imagen para el próximo movimiento
    offset_x = event.x_root
    offset_y = event.y_root

def on_image_release(event):
    global original_position_img_1, image_id, offset_x, offset_y
    # Volver a la posición original de la imagen cuando se suelta el clic
    placeholder_1.coords(image_id, original_position_img_1[0], original_position_img_1[1])


# Function to update images according to the selected Radiobutton
def update_image_source():
    update_image_from_selection()


def load_vita_images():
    """Loads all vita_tooth images and prepares them for random display."""
    global vita_images
    vita_images = []

    # Search for images in the vita_tooth directory
    for filename in os.listdir(vita_dir):
        if filename.endswith(".png"):
            image_path = os.path.join(vita_dir, filename)
            img = Image.open(image_path).convert("RGBA")  # Load with transparency
            img = img.resize((77, 112), Image.Resampling.LANCZOS)
            vita_images.append((filename, ImageTk.PhotoImage(img)))  # Save the name and image

    # Shuffle images at the start
    random.shuffle(vita_images)

def show_next_image():
    """Shows the next random image in Image 2."""
    global current_index
    if vita_images:  # Ensure the list is not empty
        current_index += 1

        # If images are finished, start over
        if current_index >= len(vita_images):  
            current_index = 0
            random.shuffle(vita_images)  # Shuffle again

        # Get the current image
        _, img_tk = vita_images[current_index]

        # Display the image on the canvas
        placeholder_2.delete("all")
        placeholder_2.create_image(50, 60, image=img_tk)
        placeholder_2.image = img_tk  # Keep reference to avoid garbage collection


def validate_first_column():
    """Validates that the Comboboxes and Sliders in the first column are filled."""
    all_valid = True

    # Validate that all Comboboxes have a value different from the initial
    for combo in first_column_comboboxes:
        if combo.get() == " ":  # Default value
            all_valid = False
            break

    # Validate that all Sliders have a value greater than 0
    if all_valid:  # Only continue if the Comboboxes are valid
        for scale in first_column_scales:
            if scale.get() == 0.0:  # Slider at its initial position
                all_valid = False
                break

    # Enable or disable the "Next" button based on validation
    next_button.config(state="normal" if all_valid else "disabled")


def reset_all_inputs():
    """Resets all Comboboxes and Sliders."""
    # Reset all Comboboxes
    for combo in all_comboboxes:
        combo.set(" ")  # Initial value

    # Reset all Sliders
    for scale in all_scales:
        scale.set(0)  # Initial value

def show_next_image():
    """Shows the next random image in Image 2 and resets the values."""
    global current_index

    # Save results of the current case
    if current_index >= 0:
        current_tooth = vita_images[current_index][0].split(".")[0]  # Get the tooth name
        update_results_matrix(current_tooth)

    # Move to the next index
    current_index += 1

    if current_index < len(vita_images):  # If there are still images available
        _, img_tk = vita_images[current_index]

        # Display the image on the canvas
        placeholder_2.delete("all")
        placeholder_2.create_image(50, 60, image=img_tk)
        placeholder_2.image = img_tk  # Keep reference to avoid garbage collection

        # Reset all Comboboxes and Sliders
        reset_all_inputs()

        # Disable the "Next" button if the first column is not filled
        next_button.config(state="disabled")
        validate_first_column()  # Automatically validate if the first column is filled

    else:  
        # All images have been shown
        save_results_to_excel()
        next_button.config(state="disabled")  # Disable "Next"
        reset_button.config(state="normal")  # Enable "Reset all"


def reset_cycle():
    """Resets the image cycle and controls."""
    global current_index
    global user_name

    # Save results to Excel before resetting
    initialize_results_matrix()

    # Ask for the name again
    ask_user_name()

    # Shuffle images again
    random.shuffle(vita_images)  
    current_index = -1  # Reset index
    placeholder_2.delete("all")  # Clear the second image canvas
    next_button.config(state="normal")  # Enable the "Next" button
    reset_button.config(state="disabled")  # Disable the "Reset all" button
    show_next_image()


# Initialize results matrix with empty rows
def initialize_results_matrix():
    global results_matrix
    results_matrix = [[None] * 6 for _ in color_labels]  # 6 columns for each tooth


def update_results_matrix(diente):
    """Updates the results matrix with the values from dropdowns and sliders."""
    global results_matrix

    # Find the row corresponding to the current tooth
    try:
        row_index = color_labels.index(diente)
    except ValueError:
        messagebox.showerror("Error", f"Tooth '{diente}' not found in the list.")
        return

    # Extract values from the dropdowns and sliders
    upper_value = first_column_comboboxes[0].get()
    upper_confidence = first_column_scales[0].get()

    central_value = first_column_comboboxes[1].get()
    central_confidence = first_column_scales[1].get()

    lower_value = first_column_comboboxes[2].get()
    lower_confidence = first_column_scales[2].get()

    # Update the corresponding row
    results_matrix[row_index] = [
        upper_value, upper_confidence,
        central_value, central_confidence,
        lower_value, lower_confidence
    ]


def save_results_to_excel():
    """Saves the results matrix to an Excel file."""
    global user_name

    # File and sheet
    file_name = "Results.xlsx"
    sheet_name = user_name

    # Create file if it doesn't exist
    if not os.path.exists(file_name):
        wb = Workbook()
        wb.save(file_name)

    # Load the existing file
    wb = load_workbook(file_name)

    # If the sheet already exists, add an incremental suffix
    original_sheet_name = sheet_name
    counter = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{original_sheet_name}_{counter}"
        counter += 1

    # Create a new sheet
    ws = wb.create_sheet(title=sheet_name)

    # Add headers
    headers = [
        "Tooth",
        "Upper Value", "Upper Confidence",
        "Central Value", "Central Confidence",
        "Lower Value", "Lower Confidence"
    ]
    ws.append(headers)

    # Write the data from the matrix to Excel
    for tooth_label, row_data in zip(color_labels, results_matrix):
        ws.append([tooth_label] + row_data)

    # Save the file
    wb.save(file_name)
    messagebox.showinfo("Success", f"Results saved to {file_name}.")







# Create the main frame for the colors
frame = tk.Frame(root)
frame.pack(pady=10, padx=10)

# Global variable to store the user's name
user_name = ""
results_matrix = [] 
original_position_img_1 = (50, 60)
image_id = None

# Ask for the user's ID at the start
ask_user_name()

# Variable to store the selected option
image_source = tk.StringVar(value="tooth")  # By default, use 'Vita Tooth'

# Create the images and assign checkboxes
color_labels = ["A1", "A2", "A3", "A3_5", "A4", "B1", "B2", "B3", "B4", "C1", "C2", "C3", "C4", "D2", "D4"]
image_dir = os.path.join(os.getcwd(), "Datasets", "reference_colors")  # Initial image directory
vita_dir = os.path.join(os.getcwd(), "Datasets", "vita_tooth")  # Image directory for 'Image 2'
image_files = [f"{label}.png" for label in color_labels]

# Create matrix to save results
initialize_results_matrix()

# Load images
images = [tk.PhotoImage(file=os.path.join(image_dir, image)) for image in image_files]

# Shared variable for the checkboxes
selected_color = tk.StringVar(value="")  

# Create containers for images and checkboxes
for idx, (color, img) in enumerate(zip(color_labels, images)):
    col_frame = tk.Frame(frame)  # Frame for each image and checkbox
    col_frame.grid(row=0, column=idx, padx=5, pady=5)

    # Image
    img_label = tk.Label(col_frame, image=img)
    img_label.pack()

    # Radiobutton under the image
    chk = tk.Radiobutton(
        col_frame, 
        text=color, 
        variable=selected_color, 
        value=color, 
        command=update_image_from_selection  
    )
    chk.pack()

# Create a main frame for organizing the center and right sections
main_center_frame = tk.Frame(root)
main_center_frame.pack(pady=20, padx=20)

# Create the area for central images
center_frame = tk.Frame(main_center_frame)
center_frame.grid(row=0, column=0, padx=10)

# Reduce top spacers to bring down the images
tk.Label(center_frame, height=7).grid(row=0, column=0, columnspan=2)  

# Placeholder for "Image 1"
placeholder_1 = tk.Canvas(center_frame, width=100, height=120, bg="SystemButtonFace", highlightthickness=0)
placeholder_1.grid(row=1, column=0, padx=10, pady=10)

# Placeholder for "Image 2" 
placeholder_2 = tk.Canvas(center_frame, width=100, height=120, bg="SystemButtonFace", highlightthickness=0)
placeholder_2.grid(row=1, column=1, padx=10, pady=10)

# Buttons and checkboxes under the central images
buttons_frame = tk.Frame(center_frame)
buttons_frame.grid(row=2, column=0, columnspan=2, pady=10)

next_button = tk.Button(buttons_frame, text="Next", command=show_next_image)
next_button.grid(row=0, column=0, columnspan=2, pady=5)

# "Reset all" button to restart the cycle
reset_button = tk.Button(buttons_frame, text="Reset all", command=reset_cycle, state="disabled")
reset_button.grid(row=2, column=0, columnspan=2, pady=5)

# Create the sliders with comboboxes to the right of the central images
sliders_frame = tk.Frame(main_center_frame)
sliders_frame.grid(row=0, column=1, padx=20)

# Create global lists to store references to the Comboboxes and Sliders
all_comboboxes = []  # All Comboboxes
all_scales = []      # All Sliders

# Create specific lists for the first column (for validation)
first_column_comboboxes = []
first_column_scales = []

# Row names
row_names = ["Upper Tooth", "Central Tooth", "Lower Tooth"]
for row_idx, row_name in enumerate(row_names):
    # Label for each row
    row_label = tk.Label(sliders_frame, text=row_name, font=("Arial", 10, "bold"))
    row_label.grid(row=row_idx, column=0, padx=10, pady=5)

    # Add sliders and comboboxes in corresponding columns
    for col_idx in range(3):  # Three columns of sliders and comboboxes per row
        slider_frame = tk.Frame(sliders_frame)
        slider_frame.grid(row=row_idx, column=col_idx + 1, padx=10, pady=5)

        # Combobox
        combo = ttk.Combobox(slider_frame, values=color_labels, width=4, state="readonly")
        combo.set(" ")  # Initial value
        combo.pack()
        all_comboboxes.append(combo)  # Save reference globally

        if col_idx == 0:  # First column
            first_column_comboboxes.append(combo)

        # Slider
        scale = tk.Scale(slider_frame, from_=0, to=1, resolution=0.1, orient=tk.HORIZONTAL)
        scale.pack()
        all_scales.append(scale)  # Save reference globally

        if col_idx == 0:  # First column
            first_column_scales.append(scale)

# Disable the "Next" button at the start
next_button.config(state="disabled")

# Bind validation to events
for combo in first_column_comboboxes:
    combo.bind("<<ComboboxSelected>>", lambda e: validate_first_column())

for scale in first_column_scales:
    scale.bind("<ButtonRelease-1>", lambda e: validate_first_column())
    scale.bind("<Motion>", lambda e: validate_first_column())

vita_images = []
current_index = -1


# Create Radiobuttons to select the reference type
reference_options_frame = tk.Frame(buttons_frame)
reference_options_frame.grid(row=1, column=0, columnspan=2, pady=10)

# When selected, image_source will be "tooth"
tk.Radiobutton(
    reference_options_frame,
    text="Reference Tooth",
    variable=image_source,
    value="tooth",  
    command=update_image_source  
).grid(row=0, column=0, padx=10)

# When selected, image_source will be "color"
tk.Radiobutton(
    reference_options_frame,
    text="Reference Color",
    variable=image_source,
    value="color",  
    command=update_image_source  
).grid(row=0, column=1, padx=10)

# Load the images at the start
load_vita_images()
update_image_from_selection()

# Show the first random image
show_next_image()

# Run the interface
root.mainloop()
