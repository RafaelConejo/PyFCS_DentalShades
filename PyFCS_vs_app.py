import os
import sys
import random
import tkinter as tk
from openpyxl import Workbook, load_workbook
from tkinter import ttk, simpledialog, messagebox
from PIL import Image, ImageTk
import pandas as pd
import ast
import time

current_dir = os.path.dirname(__file__)
sys.path.append(current_dir)


############################################################ FUNTIONS ############################################################

def exit_fullscreen(event):
    root.attributes('-fullscreen', False)


def parse_column(cell):
    try:
        # Usar `ast.literal_eval` para evaluar la cadena de forma segura
        return ast.literal_eval(cell) if isinstance(cell, str) else cell
    except Exception as e:
        print(f"Error al parsear: {cell}. Error: {e}")
        return []


def update_counter_label():
    """Updates the counter label with the current file number."""
    counter_label.config(text=f"{current_file_index}/{total_files}")

def next_file():
    """Moves to the next file, if possible."""
    global current_file_index
    if current_file_index < total_files:
        current_file_index += 1
        update_counter_label()

def previous_file():
    """Moves to the previous file, if possible."""
    global current_file_index
    if current_file_index > 0:
        current_file_index -= 1
        update_counter_label()


def ask_user_name():
    """Asks the user to enter a name."""
    global user_name
    user_name = simpledialog.askstring("ID", "Enter user ID:")
    if not user_name:  # If the user doesn't enter anything
        messagebox.showerror("Error", "You must enter a name to continue.")
        ask_user_name()  # Ask for the name again


def update_image_from_selection():
    """Updates the canvas with an image based on the selected option."""
    global image_id1
    selected_label = selected_color.get()

    if not selected_label:  # If no selection, clear the image
        shared_canvas.delete("image1")
        return

    # Set directory and image size based on the source
    directory = "Datasets/vita_tooth_test" if image_source.get() == "tooth" else "Datasets/reference_colors"
    new_size = (62, 90) if directory == "Datasets/vita_tooth_test" else (50, 50)
    image_path = os.path.join(directory, f"{selected_label}.png")

    if os.path.exists(image_path):  # Load and display image if it exists
        img = Image.open(image_path).convert("RGBA").resize((new_size), Image.Resampling.LANCZOS)
        img_tk = ImageTk.PhotoImage(img)
        shared_canvas.delete("image1")  # Clear previous image
        shared_canvas.image1 = img_tk  # Save reference
        image_id1 = shared_canvas.create_image(*coords_img1, image=img_tk, tag="image1")

        # Bind events for image manipulation
        shared_canvas.bind("<ButtonPress-1>", on_image_press)
        shared_canvas.bind("<B1-Motion>", on_image_drag)
        shared_canvas.bind("<ButtonRelease-1>", on_image_release)

    else:  # Clear image if file is missing
        shared_canvas.delete("image1")
        shared_canvas.create_text(*coords_img1, text=" ", tag="image1")


def on_image_press(event):
    """Saves the initial click coordinates for dragging the image."""
    global offset_x, offset_y
    offset_x, offset_y = event.x, event.y

def on_image_drag(event):
    """Moves the image on the canvas while dragging."""
    global offset_x, offset_y
    canvas_items = shared_canvas.find_closest(event.x, event.y)
    if canvas_items and canvas_items[0] == image_id1:
        dx, dy = event.x - offset_x, event.y - offset_y
        shared_canvas.move(image_id1, dx, dy)
        offset_x, offset_y = event.x, event.y

def on_image_release(event):
    """Resets the image to its original position after release."""
    global original_position_left_image, image_id1
    shared_canvas.coords(image_id1, *original_position_left_image)

def update_image_source():
    """Updates the displayed image when the source changes."""
    update_image_from_selection()


def load_vita_images():
    """Loads all vita_tooth images and prepares them for random display."""
    global vita_images
    vita_images = []

    # Search for images in the vita_tooth directory
    for filename in os.listdir(vita_dir):
        if filename.endswith(".png"):
            image_path = os.path.join(vita_dir, filename)
            img = Image.open(image_path).convert("RGBA")  # Load with transparency
            img = img.resize((62, 90), Image.Resampling.LANCZOS)
            vita_images.append((filename, ImageTk.PhotoImage(img)))  # Save the name and image

    # Shuffle images at the start
    random.shuffle(vita_images)        


def next_file_timer():
    """
    Returns the elapsed time since the last call and resets the timer.
    If it's the first call, it returns 0.
    """
    global start_time
    if start_time is None:  # First call
        elapsed_time = 0
    else:  # Calculate elapsed time
        elapsed_time = time.time() - start_time
        start_time = time.time()  # Restart the timer
    
    return elapsed_time


def load_image_for_tooth(tooth_name, img_frame):
    """
    Loads and displays the image for a given tooth.
    If the image exists, it is resized and displayed in the provided frame.
    """
    global image_dir
    image_path = os.path.join(image_dir, f"{tooth_name}.png")

    # Check if the image exists
    if os.path.exists(image_path):
        try:
            # Load and resize the image
            img = Image.open(image_path).convert("RGBA")
            img = img.resize((20, 20), Image.Resampling.LANCZOS)
            img_tk = ImageTk.PhotoImage(img)

            # Display the image in the specified frame
            label = tk.Label(image_frames[img_frame], image=img_tk)
            label.image = img_tk  # Keep a reference to prevent garbage collection
            label.pack(expand=True)  # Adjust position in the frame
        except Exception as e:
            print(f"Error loading image {image_path}: {e}")
            return None
    else:
        print(f"Warning: Image for {tooth_name} does not exist in {image_dir}.")
        return None


def show_next_image():
    """
    Displays the next random image in Image 2 and resets values for the next case.
    Saves results of the current case, updates the UI, and handles image display.
    """
    global current_index, results_matrix, time_matrix

    # Save results of the current case
    if current_index >= 0:
        current_tooth = vita_images[current_index][0].split(".")[0]  # Get tooth name
        update_results_matrix(current_tooth)
        update_comments_matrix(current_tooth)

        current_time = next_file_timer()
        row_index = color_labels.index(current_tooth)
        time_matrix[row_index] = current_time

    reset_all_inputs()  # Reset input fields
    current_index += 1  # Move to the next image index
    next_file()  # Proceed to next file

    # Check if there are more images
    if current_index < len(vita_images):
        _, img_tk = vita_images[current_index]

        # Display the image on the canvas
        shared_canvas.delete("image2")
        shared_canvas.image2 = img_tk
        shared_canvas.create_image(*coords_img2, image=img_tk, tag="image2")

        # Load and display saved values for the current tooth
        current_tooth = vita_images[current_index][0].split(".")[0]
        current_idx_t = color_labels.index(current_tooth)
        current_row = data.iloc[current_idx_t]
        current_row_results = results_matrix[current_idx_t]

        # Display values for "top", "middle", and "bottom"
        for section, section_values, start_idx in [
            ("top", current_row['top'], 0),
            ("middle", current_row['middle'], 3),
            ("bottom", current_row['bottom'], 6)
        ]:
            for idx, (label, value) in enumerate(section_values):
                idx_in_matrix = start_idx + idx
                if value <= 0.1:
                    all_static_texts[idx_in_matrix].config(text="")
                    all_static_texts[idx_in_matrix].grid_forget()
                    for rb in all_radiobuttons[idx_in_matrix * 5:(idx_in_matrix + 1) * 5]:
                        rb.pack_forget()
                else:
                    all_static_texts[idx_in_matrix].config(text=f"{label} -> {value}")
                    load_image_for_tooth(label, idx_in_matrix)

            # Hide remaining elements if the list is shorter than 3
            for idx in range(len(section_values), 3):
                idx_in_matrix = start_idx + idx
                all_static_texts[idx_in_matrix].config(text="")
                all_static_texts[idx_in_matrix].grid_forget()
                for rb in all_radiobuttons[idx_in_matrix * 5:(idx_in_matrix + 1) * 5]:
                    rb.pack_forget()

        # Restore previous values if any
        if current_row_results and any(value is not None for value in current_row_results):
            restore_previous_values(current_tooth)

        # Update button states
        prev_button.config(state="normal" if current_index > 0 else "disabled")
        next_button.config(state="normal" if current_index < len(vita_images) - 1 else "disabled")
        update_image_from_selection()
        validate_visible()

    else:
        # All images have been shown, finalize if the user agrees
        finalize = messagebox.askyesno("Finalize", "Your selections will be saved. Do you want to finalize?")
        if finalize:
            save_results_to_excel()
            save_time_to_excel()

            prev_button.config(state="disabled")
            next_button.config(state="disabled")  # Disable "Next"
            reset_button.config(state="normal")  # Enable "Reset all"
        else:
            current_index -= 1  # Return to the last image index


def show_previous_image():
    """
    Displays the previous image and restores its values. It saves current results,
    moves to the previous index, and updates the UI with relevant data.
    """
    global current_index, results_matrix

    if current_index > 0:  # Prevent going to negative indices
        # Save current results
        current_tooth = vita_images[current_index][0].split(".")[0]
        current_idx = color_labels.index(current_tooth)
        
        # Ensure all values are valid before updating
        if all(value is not None for value in results_matrix[current_idx]):
            update_results_matrix(current_tooth)
            update_comments_matrix(current_tooth)

        # Move to the previous image
        reset_all_inputs()
        current_index -= 1

        # Display the previous image
        _, img_tk = vita_images[current_index]
        shared_canvas.delete("image2")
        shared_canvas.image2 = img_tk
        shared_canvas.create_image(*coords_img2, image=img_tk, tag="image2")

        # Restore saved values for the previous image
        current_tooth = vita_images[current_index][0].split(".")[0]
        current_idx_t = color_labels.index(current_tooth)
        current_row = data.iloc[current_idx_t]

        # Process "top", "middle", "bottom" values
        for section, section_values, start_idx in [
            ("top", current_row['top'], 0),
            ("middle", current_row['middle'], 3),
            ("bottom", current_row['bottom'], 6)
        ]:
            for idx, (label, value) in enumerate(section_values):
                idx_in_matrix = start_idx + idx
                if value <= 0.1:
                    all_static_texts[idx_in_matrix].config(text="")
                    all_static_texts[idx_in_matrix].grid_forget()
                    for rb in all_radiobuttons[idx_in_matrix * 5:(idx_in_matrix + 1) * 5]:
                        rb.pack_forget()
                else:
                    all_static_texts[idx_in_matrix].config(text=f"{label} -> {value}")
                    load_image_for_tooth(label, idx_in_matrix)

            # Hide extra items if there are fewer than 3 values
            for idx in range(len(section_values), 3):
                idx_in_matrix = start_idx + idx
                all_static_texts[idx_in_matrix].config(text="")
                all_static_texts[idx_in_matrix].grid_forget()
                for rb in all_radiobuttons[idx_in_matrix * 5:(idx_in_matrix + 1) * 5]:
                    rb.pack_forget()

        restore_previous_values(current_tooth)
        update_image_from_selection()

    # Update button states
    prev_button.config(state="normal" if current_index > 0 else "disabled")
    next_button.config(state="normal" if current_index < len(vita_images) - 1 else "disabled")
    previous_file()  # Proceed to the previous file


def restore_previous_values(current_tooth):
    """
    Restores the values for the current tooth's radiobuttons and texts, 
    or resets them if no previous values exist.
    """
    global results_matrix, comments_matrix

    try:
        # Find the index of the current tooth in the color_labels list
        current_index = color_labels.index(current_tooth)
    except ValueError:
        messagebox.showerror("Error", f"Tooth '{current_tooth}' not found in the list.")
        return

    # Retrieve the corresponding row for the current tooth from the data
    current_row = data.iloc[current_index]
    current_comment_row = comments_matrix[current_index]

    top_values = current_row['top']    
    middle_values = current_row['middle']
    bottom_values = current_row['bottom']

    # Retrieve the result matrix for the current tooth
    current_row = results_matrix[current_index]

    # Check if the row contains non-null values
    if current_row and any(value is not None for value in current_row):
        # Iterate through the 9 columns (3 for top, 3 for middle, 3 for bottom)
        for col_idx in range(9):  # 9 columns per row (3 for top, 3 for middle, 3 for bottom)
            value = current_row[col_idx]

            # If there's a value to restore
            if value != -1:
                # Update the corresponding text label
                if col_idx < 3:
                    all_static_texts[col_idx].config(text=f"{top_values[col_idx][0]} -> {top_values[col_idx][1]}")
                elif col_idx >= 3 and col_idx < 6:
                    all_static_texts[col_idx].config(text=f"{middle_values[col_idx - 3][0]} -> {middle_values[col_idx - 3][1]}")
                else:
                    all_static_texts[col_idx].config(text=f"{bottom_values[col_idx - 6][0]} -> {bottom_values[col_idx - 6][1]}")

                # Restore the radiobutton value
                radiobutton_values[col_idx].set(value)

                # Ensure the radiobuttons are visible
                for rb in all_radiobuttons[col_idx * 5: (col_idx + 1) * 5]:  # 5 buttons per row
                    rb.pack(side=tk.LEFT)

        # Restore the comments for the current row in the text entries
        for col_comment in range(3):
            comment = current_comment_row[col_comment]
            additional_text_entries[col_comment].delete(0, tk.END)  # Clear previous text
            additional_text_entries[col_comment].insert(0, comment if comment else "")

    else:
        # If the row is empty or contains only None, reset all inputs
        reset_all_inputs()


def clear_comments():
    """Clears the text entries and resets the comments matrix."""
    global comments_matrix

    for entry in additional_text_entries:
        if isinstance(entry, tk.Entry):  # Ensure the entry is an instance of tk.Entry
            entry.delete(0, tk.END)  # Clear the content of the entry
        else:
            print(f"Warning: {entry} is not a tk.Entry") 


def validate_visible():
    """Validates that at least one Radiobutton in each visible row is selected."""
    all_valid = True  # Initially assume all are valid

    # Iterate over the radiobutton value variables
    for idx, rb_value in enumerate(radiobutton_values):
        # If the associated label text is not hidden, check the radiobutton value
        if all_static_texts[idx].cget("text") != "":   # Check if the label is visible
            if rb_value.get() == -1:  # No radiobutton selected
                all_valid = False
                break

    # Enable or disable the "Next" button based on the validation
    next_button.config(state="normal" if all_valid else "disabled")


def reset_all_inputs():
    """Resets all Radiobuttons to their initial state (value -1) and restores visibility."""
    clear_comments()

    # Reset all radiobutton values to -1 (deselect them)
    for rb_value in radiobutton_values:
        rb_value.set(-1)  # Reset to initial value

    # Iterate over all static text labels and radiobuttons
    for idx in range(len(all_static_texts)):
        # Check if the text is empty (indicating that the row is hidden)
        if all_static_texts[idx].cget("text") == "":  # If the text is empty, restore visibility
            # Make radiobuttons visible if they were hidden
            for rb in all_radiobuttons[idx * 5: (idx + 1) * 5]:  # Considering there are 5 buttons per row
                if not rb.winfo_ismapped():  # If the radiobutton is hidden
                    rb.pack(side=tk.LEFT)  # Make it visible again
            # Restore text if needed
            all_static_texts[idx].config(text="")  # You can set the text you want here

        # Deselect all radiobuttons
        for rb in all_radiobuttons[idx * 5: (idx + 1) * 5]:
            rb.deselect()  # Deselect the radiobuttons

    # Destroy all widgets inside each frame (including images)
    for frame in image_frames:
        for widget in frame.winfo_children():
            widget.destroy()


def reset_cycle():
    """Resets the image cycle and controls."""
    global current_index
    global user_name
    global current_file_index
    global start_time

    current_file_index = 0

    # Save results to Excel before resetting
    initialize_results_matrix()

    # Ask for the name again
    ask_user_name()
    start_time = time.time()

    # Shuffle images again
    random.shuffle(vita_images)  
    current_index = -1  # Reset index
    shared_canvas.delete("image2")  # Clear the second image canvas
    next_button.config(state="normal")  # Enable the "Next" button
    reset_button.config(state="disabled")  # Disable the "Reset all" button
    show_next_image()
    prev_button.config(state="disabled")


# Initialize results matrix with empty rows
def initialize_results_matrix():
    global results_matrix
    results_matrix = [[None] * 9 for _ in color_labels]  # 9 columns for each tooth

def initialize_comments_matrix():
    global comments_matrix
    comments_matrix = [[None] * 3 for _ in color_labels]  # 3 columns for each tooth

def initialize_time_matrix():
    global time_matrix
    time_matrix = [None for _ in color_labels]


def update_results_matrix(diente):
    """Updates the results matrix with the values from dropdowns and sliders."""
    global results_matrix

    try:
        row_idx = color_labels.index(diente)
    except ValueError:
        messagebox.showerror("Error", f"Tooth '{diente}' not found.")
        return

    for col_idx in range(9):  # 3 columns: top, middle, bottom
        # Check if the set of radiobuttons is visible
        if results_matrix[row_idx][col_idx] != -1:  # If the radiobuttons are visible
            # Get the selected value from the radiobuttons
            selected_value = radiobutton_values[col_idx].get()
            results_matrix[row_idx][col_idx] = selected_value


def update_comments_matrix(diente):
    """Updates the comments matrix with the values from the text entries for the given tooth."""
    global comments_matrix

    try:
        # Get the row index of the tooth in the matrix
        row_idx = color_labels.index(diente)
    except ValueError:
        messagebox.showerror("Error", f"Tooth '{diente}' not found in the list.")
        return

    # Loop through the text entries for the current tooth
    for col_idx, entry in enumerate(additional_text_entries):
        # Get and clean the text from the entry
        comment = entry.get().strip()
        # Save the comment in the matrix, or `None` if empty
        comments_matrix[row_idx][col_idx] = comment if comment else None


def save_results_to_excel():
    """Saves the results matrix to an Excel file."""
    global user_name, comments_matrix, results_matrix

    # File and sheet
    file_name = "Results\PyFCS_Val_Results.xlsx"
    sheet_name = user_name

    # Create file if it doesn't exist
    if not os.path.exists(file_name):
        wb = Workbook()
        wb.save(file_name)

    # Load the existing file
    wb = load_workbook(file_name)

    # If the sheet already exists, add an incremental suffix
    original_sheet_name = sheet_name
    counter = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{original_sheet_name}_{counter}"
        counter += 1

    # Create a new sheet
    ws = wb.create_sheet(title=sheet_name)

    # Add headers
    headers = [
        "Tooth",
        "UP_1", "UP_2", "UP_3", "U_Comment", 
        "CP_1", "CP_2", "CP_3", "C_Comment", 
        "LP_1", "LP_2", "LP_3", "L_Comment"
    ]
    ws.append(headers)

    # Write the data from the results_matrix to Excel
    for tooth_label, row_data, comment_row in zip(color_labels, results_matrix, comments_matrix):
            numeric_values = [value if value != -1 else "" for value in row_data]
            processed_comments = [comment if comment is not None else "" for comment in comment_row]

            full_row = numeric_values[:3] + [processed_comments[0]] + numeric_values[3:6] + [processed_comments[1]] + numeric_values[6:] + [processed_comments[2]]
            
            ws.append([tooth_label] + full_row)

    # Save the file
    wb.save(file_name)
    messagebox.showinfo("Success", f"Results saved to {file_name}.")


def save_time_to_excel():
    """Saves the results Time matrix to an Excel file."""
    global user_name, time_matrix

    # File and sheet setup
    file_name = "Results\PyFCS_Val_Time.xlsx"
    sheet_name = f"{user_name}_Time"

    # Create the file if it doesn't exist
    if not os.path.exists(file_name):
        wb = Workbook()
        wb.save(file_name)

    # Load the existing file
    wb = load_workbook(file_name)

    # If the sheet already exists, add an incremental suffix
    original_sheet_name = sheet_name
    counter = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{original_sheet_name}_{counter}"
        counter += 1

    # Create a new sheet
    ws = wb.create_sheet(title=sheet_name)

    # Add headers
    ws.append(["Tooth", "Elapsed Time (seconds)", "Elapsed Time (minutes)"])

    # Write the times to the Excel file
    total_time = 0
    for tooth_label, elapsed_time in zip(color_labels, time_matrix):
        # Replace None values with 0
        if elapsed_time is None:
            elapsed_time = 0
        elapsed_time_minutes = elapsed_time / 60  # Convert to minutes
        ws.append([tooth_label, elapsed_time, elapsed_time_minutes])
        total_time += elapsed_time

    # Add a row for the total
    total_time_minutes = total_time / 60  # Convert total time to minutes
    ws.append(["Total", total_time, total_time_minutes])

    # Save the file
    wb.save(file_name)


############################################################ Main Windown ############################################################

# Initialize the main window
root = tk.Tk()

# Set the window to fullscreen
# root.geometry("1300x800") 
root.attributes('-fullscreen', True)

# Bind the Escape key to exit fullscreen mode
root.bind('<Escape>', exit_fullscreen)

# Set the window title
root.title("Color Selector")

# File path to the Excel file containing PyFCS data
file_path = os.path.join(os.getcwd(), "Datasets", "results_opt_1.xlsx")

# Read the data from the Excel file
data = pd.read_excel(file_path)

# List of columns that need conversion
columns_to_convert = ['top', 'middle', 'bottom']

# Convert the string representations into lists of tuples for the specified columns
for col in columns_to_convert:
    data[col] = data[col].apply(ast.literal_eval)

# Iterate through each row in the DataFrame to extract the top, middle, and bottom values
for idx, row in data.iterrows():
    top_values = row['top']    
    middle_values = row['middle']  
    bottom_values = row['bottom']

# Initialize the timer and time matrix for tracking elapsed times
start_time = None
time_matrix = []


############################################################ INTERFACE ############################################################

# Create the main frame for the colors
frame = tk.Frame(root)
frame.pack(pady=10, padx=10)

# Global variable to store the user's name
user_name = ""
results_matrix = [] 
comments_matrix = []
original_position_left_image = (150, 200)  # left image original position
left_image_id = None  # ID left image
current_file_index = 0

# Ask for the user's ID at the start
ask_user_name()
start_time = time.time()

# Variable to store the selected option
image_source = tk.StringVar(value="tooth")  # By default, use 'Vita Tooth'

# Create the images and assign checkboxes
color_labels = ["A1", "A2", "A3", "A3_5", "A4", "B1", "B2", "B3", "B4", "C1", "C2", "C3", "C4", "D2", "D3", "D4"]
image_dir = os.path.join(os.getcwd(), "Datasets", "reference_colors")  # Initial image directory
vita_dir = os.path.join(os.getcwd(), "Datasets", "vita_tooth")  # Image directory for 'Image 2'
image_files = [f"{label}.png" for label in color_labels]

total_files = len([f for f in os.listdir(vita_dir) if os.path.isfile(os.path.join(vita_dir, f))])

# Create matrix to save results
initialize_results_matrix()
initialize_comments_matrix()
initialize_time_matrix()

# Load images
images = [tk.PhotoImage(file=os.path.join(image_dir, image)) for image in image_files]

# Shared variable for the checkboxes
selected_color = tk.StringVar(value=0)  

# Create containers for images and checkboxes
for idx, (color, img) in enumerate(zip(color_labels, images)):
    col_frame = tk.Frame(frame)  # Frame for each image and checkbox
    col_frame.grid(row=0, column=idx, padx=5, pady=5)

    # Image
    img_label = tk.Label(col_frame, image=img)
    img_label.pack()

    # Radiobutton under the image
    chk = tk.Radiobutton(
        col_frame, 
        text=color, 
        variable=selected_color, 
        value=color, 
        command=update_image_from_selection  
    )
    chk.pack()

# Create a main frame for organizing the center and right sections
main_center_frame = tk.Frame(root)
main_center_frame.pack(pady=20, padx=20)

# Create the area for central images
center_frame = tk.Frame(main_center_frame)
center_frame.grid(row=0, column=0, padx=10)

# Reduce top spacers to bring down the images
tk.Label(center_frame, height=0).grid(row=0, column=0, columnspan=2)  

# Create canvas for two images
shared_canvas = tk.Canvas(center_frame, width=440, height=400, bg="SystemButtonFace", highlightthickness=0)
shared_canvas.grid(row=1, column=0, padx=10, pady=1, columnspan=2)

# Coordenates of the two images 
coords_img1 = original_position_left_image
coords_img2 = (300, 200)  

# Buttons and checkboxes under the central images
buttons_frame = tk.Frame(center_frame)
buttons_frame.grid(row=3, column=0, columnspan=2, pady=2)

counter_label = tk.Label(buttons_frame, text="1/16", font=("Arial", 10))
counter_label.grid(row=0, column=0, columnspan=2, pady=5)

prev_button = tk.Button(buttons_frame, text="Previous", command=show_previous_image, state="disabled")
prev_button.grid(row=1, column=0, padx=10)

next_button = tk.Button(buttons_frame, text="Next", command=show_next_image)
next_button.grid(row=1, column=1, padx=10)

# "Reset all" button to restart the cycle
reset_button = tk.Button(buttons_frame, text="Reset all", command=reset_cycle, state="disabled")
reset_button.grid(row=3, column=0, columnspan=2, pady=5)

# Create the sliders with comboboxes to the right of the central images
sliders_frame = tk.Frame(main_center_frame)
sliders_frame.grid(row=0, column=1, padx=10, pady=1)

# Create global lists to hold references to various widgets
all_static_texts = []  # References to static texts
all_radiobuttons = []  # References to radiobuttons
radiobutton_values = []  # Stores the selected values from the radiobuttons
image_frames = []  # References to the image widget frames

# Names for the rows in the interface
row_names = ["Upper Tooth", "Central Tooth", "Lower Tooth"]

# Loop through each row name and create the corresponding UI elements
for row_idx, row_name in enumerate(row_names):
    # Label for each row
    row_label = tk.Label(sliders_frame, text=row_name, font=("Arial", 10, "bold"))
    row_label.grid(row=row_idx, column=0, padx=10, pady=0)

    # Add static text, images, and radiobuttons for each column in the row
    for col_idx in range(3):  # Three columns per row
        control_frame = tk.Frame(sliders_frame)
        control_frame.grid(row=row_idx, column=col_idx + 1, padx=10, pady=7)

        # Empty frame for the image 
        image_frame = tk.Frame(control_frame, width=20, height=20)
        image_frame.pack(padx=5)
        image_frame.pack_propagate(False)   # Prevent the frame from resizing automatically
        image_frames.append(image_frame)   # Save reference to image frame

        # Static text label
        static_text = tk.Label(control_frame, text="Text", font=("Arial", 10))
        static_text.pack() 
        all_static_texts.append(static_text)  # Save reference to the static text

        # Frame for radiobuttons
        radiobutton_frame = tk.Frame(control_frame)
        radiobutton_frame.pack(side="bottom", pady=5)

        # Variable for the radiobuttons' selected value
        rb_value = tk.IntVar(value=-1)  # Default value for the radiobuttons
        radiobutton_values.append(rb_value)  # Store the radiobutton value variable

        # Create 5 radiobuttons (values from 1 to 5)
        for rb_idx in range(1, 6):  
            rb = tk.Radiobutton(
                radiobutton_frame, 
                text=str(rb_idx),      # Text on the radiobutton
                variable=rb_value, 
                value=rb_idx,          # Value for the radiobutton
                indicatoron=False,     # Makes the button square with aligned text
                width=2,               # Adjust width for better appearance
                command=validate_visible  # Command to validate the visibility when selected
            )
            rb.pack(side=tk.LEFT)  # Pack the radiobutton horizontally
            all_radiobuttons.append(rb)  # Save reference to the radiobutton

# Create a new frame for the additional sliders to the right
additional_text_frame = tk.Frame(main_center_frame)
additional_text_frame.grid(row=0, column=2, padx=10, pady=1)

# Add new column 
header_label = tk.Label(additional_text_frame, text="Comments", font=("Arial", 12, "bold"))
header_label.grid(row=0, column=0, padx=0, pady=0, columnspan=2)

# List to add text
additional_text_entries = []
for row_idx in range(len(row_names)):
    text_entry = tk.Entry(additional_text_frame, width=25, font=("Arial", 10))
    text_entry.grid(row=row_idx + 1, column=0, padx=10, pady=20)  

    additional_text_entries.append(text_entry)

vita_images = []
current_index = -1


# Create Radiobuttons to select the reference type
reference_options_frame = tk.Frame(buttons_frame)
reference_options_frame.grid(row=2, column=0, columnspan=2, pady=10)

# When selected, image_source will be "tooth"
tk.Radiobutton(
    reference_options_frame,
    text="Reference Tooth",
    variable=image_source,
    value="tooth",  
    command=update_image_source  
).grid(row=0, column=0, padx=10)

# When selected, image_source will be "color"
tk.Radiobutton(
    reference_options_frame,
    text="Reference Color",
    variable=image_source,
    value="color",  
    command=update_image_source  
).grid(row=0, column=1, padx=10)

# Load the images at the start
load_vita_images()
update_image_from_selection()

# Show the first random image
show_next_image()

next_button.config(state="disabled")
prev_button.config(state="disabled")

# Run the interface
root.mainloop()
